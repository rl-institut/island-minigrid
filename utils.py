import pandas as pd
import base64
from oemof.tools.economics import annuity
from openpyxl import load_workbook


def read_input_file(filename):
    """Parse a .xlsx input file"""

    wb = load_workbook(filename=filename)
    sheet_names = wb.sheetnames

    for sn in ("costs", "timeseries", "settings"):
        if sn not in sheet_names:
            raise ValueError(
                f"The sheet '{sn}' is missing in your input file {filename}"
            )

    name = "costs"
    headers = [c.value for c in wb[name][1]]
    df = pd.DataFrame(tuple(wb[name].values)[2:], columns=headers)
    # drop the lines which do not define a new asset
    df = df.loc[df.asset.notna()]
    # drop the columns with no headers
    df_costs = df[df.columns.dropna()].set_index("asset").fillna(0)

    for col_name in (
        "asset",
        "capex_variable",
        "opex_variable",
        "lifetime",
        "co2_emissions",
        "density",
        "energy_density",
        "volumetric_cost",
    ):
        if col_name not in headers:
            raise ValueError(
                f"The column header '{col_name}' is missing in your input file {filename} under the '{name}' sheet"
            )

    name = "timeseries"
    headers = [c.value for c in wb[name][1]]

    for hd in ("CriticalDemand", "Demand", "SolarGen"):
        if hd not in headers:
            raise ValueError(
                f"The column header '{hd}' is missing in your input file {filename} under the '{name}' sheet"
            )

    df_timeseries = pd.DataFrame(tuple(wb[name].values)[1:], columns=headers)

    name = "settings"
    headers = [c.value for c in wb[name][1]]

    for hd in ("param", "value"):
        if hd not in headers:
            raise ValueError(
                f"The column header '{hd}' is missing in your input file {filename} under the '{name}' sheet"
            )
    df = pd.DataFrame(tuple(wb[name].values)[1:], columns=headers)
    df = df[df.columns.dropna()].set_index("param")
    df_settings = df["value"]

    # compute the epc needed for oemof investments if not provided
    if "annuity" not in df_costs.columns:
        wacc = df_settings.wacc
        project_lifetime = df_costs.lifetime.project
        # correspond to equation (6)
        df_costs["annuity"] = df_costs.apply(
            lambda x: annualized_capex(
                x.capex_variable, project_lifetime, x.lifetime, wacc
            )
            + x.opex_fix,
            axis=1,
        )

    name = "sensitivity"
    headers = [c.value for c in wb[name][1]]

    for hd in ("category", "variable_name", "min_val", "max_val", "step"):
        if hd not in headers:
            raise ValueError(
                f"The column header '{hd}' is missing in your input file {filename} under the '{name}' sheet"
            )
    df_sensitivity = pd.DataFrame(tuple(wb[name].values)[2:], columns=headers)
    if df_sensitivity.empty is False:
        return df_costs, df_timeseries, df_settings, df_sensitivity
    else:
        return (
            df_costs,
            df_timeseries,
            df_settings,
        )


def encode_image_file(img_path):
    """Encode image files to load them in the dash layout under img html tag

    Parameters
    ----------
    img_path: str
        path to the image file

    Returns
    -------
    encoded_img: bytes
        encoded bytes of the image file

    """

    try:
        with open(img_path, "rb") as ifs:
            encoded_img = base64.b64encode(ifs.read())
    except FileNotFoundError:
        encoded_img = base64.b64encode(bytes())
    return encoded_img


def annualized_capex(
    investment_t0, project_lifetime, asset_lifetime=None, wacc=0.05, tax=0
):
    """Return output of capex_from_investment annualised"""
    capex = capex_from_investment(
        investment_t0, project_lifetime, asset_lifetime, wacc, tax
    )
    return annuity(capex, project_lifetime, wacc)


def capex_from_investment(
    investment_t0, project_lifetime, asset_lifetime=None, wacc=0.05, tax=0
):
    """

    Parameters
    ----------
    investment_t0: float
       Specific investment in year 0

    project_lifetime: int
        Project lifetime in years

    asset_lifetime: int
        Asset lifetime in years

    wacc: float
        Discount factor

    tax: float
        Tax factor

    Returns
    -------
    Capex of the asset

    """

    if asset_lifetime is None:
        asset_lifetime = project_lifetime

    # [quantity, investment, installation, weight, lifetime, om, first_investment]
    if project_lifetime == asset_lifetime:
        number_of_investments = 1
    else:
        number_of_investments = int(round(project_lifetime / asset_lifetime + 0.5))
    # costs with quantity and import tax at t=0
    first_time_investment = investment_t0 * (1 + tax)

    for count_of_replacements in range(0, number_of_investments):
        # Very first investment is in year 0
        if count_of_replacements == 0:
            capex = first_time_investment
        else:
            # replacements taking place in year = number_of_replacement * asset_lifetime
            if count_of_replacements * asset_lifetime != project_lifetime:
                capex = capex + first_time_investment / (
                    (1 + wacc) ** (count_of_replacements * asset_lifetime)
                )

    # # Substraction of component value at end of life with last replacement (= number_of_investments - 1)
    # if number_of_investments * asset_lifetime > project_lifetime:
    #     last_investment = first_time_investment / (
    #         (1 + wacc) ** ((number_of_investments - 1) * asset_lifetime)
    #     )
    #     linear_depreciation_last_investment = last_investment / asset_lifetime
    #     capex = capex - linear_depreciation_last_investment * (
    #         number_of_investments * asset_lifetime - project_lifetime
    #     ) / ((1 + wacc) ** (project_lifetime))

    return capex
